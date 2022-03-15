from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어 column 만 가지고 오기
col_B = ws['B'] # B열 모든 셀들의 정보
print(col_B)
for cell in col_B:
    print(cell.value)

# 영어, 수학 column 함께 가지고 오기
col_range = ws['B:C'] # B와 C열 모든 셀들의 정보
for cols in col_range: # cols : B열... C열...
    for cell in cols:
        print(cell.value)

# 1번째 row 만 가지고 오기
row_title = ws[1]
for cell in row_title:
    print(cell.value)

# 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지오 오기
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=' ')
    print()

from openpyxl.utils.cell import coordinate_from_string

# coordinate_from_string : col, row 의 위치를 튜플로 각자 보여주기
# 2번째 줄부터 마지막 줄까지 가지고 오기
row_range = ws[2:ws.max_row] 
for rows in row_range:
    for cell in rows:
        print(cell.value, end=' ')
        print(cell.coordinate, end=' ') # A10, AZ250
        xy = coordinate_from_string(cell.coordinate) # ('A', 2)('B', 2)
        print(xy, end=(''))
        print(xy[0], end='') # A
        print(xy[1], end=' ') # 1
    print()

# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[0].value)


# 전체 columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

# 전체 row
for row in ws.iter_rows():
    print(row[1].value)

# 전체 column
for column in ws.iter_cols():
    print(column[1].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어
    print(row)

for column in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(column)

wb.save("sample.xlsx")