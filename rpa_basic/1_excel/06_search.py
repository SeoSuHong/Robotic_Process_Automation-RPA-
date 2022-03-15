from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active

# 영어 점수가 80점 이상인 학생
for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) >= 80:
        print(row[0].value, "번 학생은 영어 천재")

# '영어' 과목을 '컴퓨터'로 바꾸는 작업
for row in ws.iter_rows(min_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")
