from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# # 수식 그대로 가져오고 있음 ex) =SUM(1, 2, 3)
for row in ws.values:
    for cell in row:
        print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# data_only : 수식이 아닌 실제 데이터를 가지고 옴

# 코드로 작성한 수식은 계산이 안되고 식만 들어갈 뿐.
# 계산은 excel문서가 가동되면서 계산됨
# excel문서를 열어서 저장 후 코드를 실행시키면 계산 결과가 출력
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)
